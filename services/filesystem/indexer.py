"""File content indexer."""

import os
import mimetypes
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Generator
import fnmatch

import structlog
from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams, PointStruct

from config import FileSystemServiceConfig
from models import FileMetadata, FileChunk

logger = structlog.get_logger(__name__)


class FileIndexer:
    """Indexes files and their content."""
    
    def __init__(
        self,
        config: FileSystemServiceConfig,
        qdrant_client: QdrantClient,
        embedder
    ):
        """Initialize the file indexer."""
        self.config = config
        self.qdrant = qdrant_client
        self.embedder = embedder
        self._ensure_collection()
    
    def _ensure_collection(self):
        """Ensure the Qdrant collection exists."""
        try:
            self.qdrant.get_collection(self.config.collection_name)
            logger.info("Collection exists", collection=self.config.collection_name)
        except:
            # Create collection if it doesn't exist
            self.qdrant.create_collection(
                collection_name=self.config.collection_name,
                vectors_config=VectorParams(
                    size=384,  # all-MiniLM-L6-v2 embedding size
                    distance=Distance.COSINE
                )
            )
            logger.info("Created collection", collection=self.config.collection_name)
    
    def should_index_file(self, file_path: Path) -> bool:
        """Check if a file should be indexed."""
        # Check size
        try:
            size_mb = file_path.stat().st_size / (1024 * 1024)
            if size_mb > self.config.max_file_size_mb:
                return False
        except:
            return False
        
        # Check extension
        ext = file_path.suffix.lower()
        if ext not in self.config.text_extensions and ext not in self.config.document_extensions:
            return False
        
        # Check exclusion patterns
        file_str = str(file_path)
        for pattern in self.config.exclude_patterns:
            if fnmatch.fnmatch(file_str, pattern):
                return False
        
        return True
    
    def extract_text(self, file_path: Path) -> Optional[str]:
        """Extract text content from a file."""
        ext = file_path.suffix.lower()
        
        try:
            # Plain text files
            if ext in self.config.text_extensions:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    return f.read()
            
            # PDF files
            elif ext == '.pdf':
                try:
                    import pypdf
                    with open(file_path, 'rb') as f:
                        reader = pypdf.PdfReader(f)
                        text = []
                        for page in reader.pages:
                            text.append(page.extract_text())
                        return '\n\n'.join(text)
                except ImportError:
                    logger.warning("pypdf not installed, skipping PDF", file=str(file_path))
                    return None
            
            # DOCX files
            elif ext == '.docx':
                try:
                    import docx
                    doc = docx.Document(file_path)
                    return '\n\n'.join([para.text for para in doc.paragraphs])
                except ImportError:
                    logger.warning("python-docx not installed, skipping DOCX", file=str(file_path))
                    return None
            
            # ODT files (LibreOffice Writer)
            elif ext == '.odt':
                try:
                    from odf import text, teletype
                    from odf.opendocument import load
                    doc = load(file_path)
                    paragraphs = []
                    for paragraph in doc.getElementsByType(text.P):
                        paragraphs.append(teletype.extractText(paragraph))
                    return '\n\n'.join(paragraphs)
                except ImportError:
                    logger.warning("odfpy not installed, skipping ODT", file=str(file_path))
                    return None
            
            # ODS files (LibreOffice Calc)
            elif ext == '.ods':
                try:
                    from odf.opendocument import load
                    from odf.table import Table, TableRow, TableCell
                    from odf import teletype
                    doc = load(file_path)
                    text_parts = []
                    for table in doc.getElementsByType(Table):
                        for row in table.getElementsByType(TableRow):
                            row_data = []
                            for cell in row.getElementsByType(TableCell):
                                cell_text = teletype.extractText(cell)
                                if cell_text:
                                    row_data.append(cell_text)
                            if row_data:
                                text_parts.append('\t'.join(row_data))
                    return '\n'.join(text_parts)
                except ImportError:
                    logger.warning("odfpy not installed, skipping ODS", file=str(file_path))
                    return None
            
            # XLSX/XLS files (Excel)
            elif ext in ['.xlsx', '.xls']:
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path, read_only=True, data_only=True)
                    text_parts = []
                    for sheet in wb.worksheets:
                        text_parts.append(f"Sheet: {sheet.title}")
                        for row in sheet.iter_rows(values_only=True):
                            row_data = [str(cell) for cell in row if cell is not None]
                            if row_data:
                                text_parts.append('\t'.join(row_data))
                    return '\n'.join(text_parts)
                except ImportError:
                    logger.warning("openpyxl not installed, skipping Excel", file=str(file_path))
                    return None
            
            # CSV files
            elif ext == '.csv':
                try:
                    import csv
                    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                        reader = csv.reader(f)
                        rows = ['\t'.join(row) for row in reader]
                        return '\n'.join(rows)
                except Exception as e:
                    logger.error("Error reading CSV", file=str(file_path), error=str(e))
                    return None
            
            # Other document types
            else:
                return None
                
        except Exception as e:
            logger.error("Error extracting text", file=str(file_path), error=str(e))
            return None
    
    def chunk_text(self, text: str, file_path: Path) -> Generator[FileChunk, None, None]:
        """Split text into overlapping chunks."""
        if not text:
            return
        
        chunk_size = self.config.chunk_size
        overlap = self.config.chunk_overlap
        
        # Get file metadata
        stat = file_path.stat()
        metadata = FileMetadata(
            path=str(file_path),
            filename=file_path.name,
            extension=file_path.suffix,
            size_bytes=stat.st_size,
            modified_time=datetime.fromtimestamp(stat.st_mtime),
            created_time=datetime.fromtimestamp(stat.st_ctime),
            mime_type=mimetypes.guess_type(str(file_path))[0]
        )
        
        # Create chunks
        start = 0
        chunk_index = 0
        
        while start < len(text):
            end = min(start + chunk_size, len(text))
            chunk_text = text[start:end]
            
            if chunk_text.strip():  # Only yield non-empty chunks
                yield FileChunk(
                    file_path=str(file_path),
                    chunk_index=chunk_index,
                    content=chunk_text,
                    char_start=start,
                    char_end=end,
                    metadata=metadata
                )
                chunk_index += 1
            
            start += chunk_size - overlap
    
    def index_file(self, file_path: Path) -> int:
        """Index a single file and return number of chunks created."""
        if not self.should_index_file(file_path):
            return 0
        
        logger.debug("Indexing file", file=str(file_path))
        
        # Extract text
        text = self.extract_text(file_path)
        if not text:
            return 0
        
        # Create chunks
        chunks_indexed = 0
        points = []
        
        for chunk in self.chunk_text(text, file_path):
            # Generate embedding
            embedding = self.embedder.encode(chunk.content).tolist()
            
            # Create point for Qdrant
            point = PointStruct(
                id=abs(hash(f"{chunk.file_path}_{chunk.chunk_index}")),
                vector=embedding,
                payload={
                    "file_path": chunk.file_path,
                    "filename": chunk.metadata.filename,
                    "extension": chunk.metadata.extension,
                    "chunk_index": chunk.chunk_index,
                    "content": chunk.content,
                    "size_bytes": chunk.metadata.size_bytes,
                    "modified_time": chunk.metadata.modified_time.isoformat(),
                    "mime_type": chunk.metadata.mime_type,
                }
            )
            points.append(point)
            chunks_indexed += 1
        
        # Upload to Qdrant in batches
        if points:
            self.qdrant.upsert(
                collection_name=self.config.collection_name,
                points=points
            )
        
        return chunks_indexed
    
    def index_directory(
        self,
        directory: Path,
        recursive: bool = True
    ) -> tuple[int, int, List[str]]:
        """
        Index all files in a directory.
        Returns: (files_indexed, chunks_created, errors)
        """
        files_indexed = 0
        chunks_created = 0
        errors = []
        
        pattern = "**/*" if recursive else "*"
        
        for file_path in directory.glob(pattern):
            if not file_path.is_file():
                continue
            
            try:
                chunks = self.index_file(file_path)
                if chunks > 0:
                    files_indexed += 1
                    chunks_created += chunks
                    
            except Exception as e:
                error_msg = f"{file_path}: {str(e)}"
                errors.append(error_msg)
                logger.error("Error indexing file", file=str(file_path), error=str(e))
        
        logger.info(
            "Directory indexing complete",
            files=files_indexed,
            chunks=chunks_created,
            errors=len(errors)
        )
        
        return files_indexed, chunks_created, errors

